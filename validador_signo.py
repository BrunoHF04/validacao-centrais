#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Validador local de JSON para os sistemas CEP e CESDI da SIGNO.
Analisa campos obrigatórios, tipos de dados, formatos e cruza códigos
com as tabelas dos manuais oficiais.

Uso:
    python validador_signo.py <arquivo.json> --sistema [CEP|CESDI]
    python validador_signo.py exemplos/cep_valido.json --sistema CEP
    python validador_signo.py exemplos/cesdi_invalido.json --sistema CESDI
"""

import argparse
import io
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Garante UTF-8 na saída do console (no-op quando rodando como .exe sem console)
try:
    if sys.stdout is not None and hasattr(sys.stdout, "encoding") \
            and sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if sys.stderr is not None and hasattr(sys.stderr, "encoding") \
            and sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
except Exception:
    pass

import pandas as pd

# ─────────────────────────────────────────────
# CONSTANTES DE LOCALIZAÇÃO DOS MANUAIS
# Compatível com PyInstaller (sys._MEIPASS) e execução normal
# ─────────────────────────────────────────────
def _get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS)
    return Path(__file__).parent

BASE_DIR = _get_base_dir()
MANUAIS_CEP_DIR = BASE_DIR / "manuais" / "CEP"
MANUAIS_CESDI_DIR = BASE_DIR / "manuais" / "CESDI"

# ─────────────────────────────────────────────
# CARREGAMENTO DAS TABELAS DOS MANUAIS
# ─────────────────────────────────────────────

def carregar_tabela_csv(caminho: Path) -> dict:
    """Lê CSV com separador ';' e retorna dicionário {código: descrição}."""
    try:
        df = pd.read_csv(caminho, sep=";", encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        # Primeira coluna = código, segunda = descrição
        col_codigo = df.columns[0]
        col_desc = df.columns[1]
        df[col_codigo] = df[col_codigo].str.strip()
        df[col_desc] = df[col_desc].str.strip()
        return {row[col_codigo]: row[col_desc] for _, row in df.iterrows()
                if pd.notna(row[col_codigo]) and row[col_codigo].isdigit()}
    except Exception as e:
        print(f"  [AVISO] Não foi possível carregar {caminho}: {e}")
        return {}


def carregar_tabela_xlsx(caminho: Path, aba: str, col_codigo: int = 0, col_desc: int = 1,
                          skiprows: int = 1) -> dict:
    """Lê aba de XLSX e retorna dicionário {código: descrição}."""
    try:
        df = pd.read_excel(caminho, sheet_name=aba, dtype=str, header=skiprows)
        resultado = {}
        for _, row in df.iterrows():
            vals = list(row)
            cod = str(vals[col_codigo]).strip() if col_codigo < len(vals) else ""
            desc = str(vals[col_desc]).strip() if col_desc < len(vals) else ""
            if cod and cod not in ("nan", "None") and re.match(r"^\d+$", cod):
                resultado[cod] = desc
        return resultado
    except Exception as e:
        print(f"  [AVISO] Não foi possível carregar aba '{aba}' de {caminho}: {e}")
        return {}


def carregar_tabelas_cesdi_xlsx(caminho: Path) -> dict:
    """Carrega todas as abas relevantes do XLSX do CESDI."""
    tabelas = {}
    mapa_abas = {
        "qualidadeParte":         ("QualidadeDaParte",       1, 2, 1),
        "tipoBemEdireito":        ("TipoBensEDireito",        0, 1, 0),
        "tipoReferenciaCadastral":("TipoReferenciaCadastral", 0, 1, 0),
    }
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
    except Exception as e:
        print(f"  [AVISO] Não foi possível abrir {caminho}: {e}")
        return tabelas

    for chave, (nome_aba, col_c, col_d, skip) in mapa_abas.items():
        try:
            ws = wb[nome_aba]
            resultado = {}
            rows = list(ws.iter_rows(values_only=True))
            data_rows = rows[skip + 1:]  # pula cabeçalho
            for row in data_rows:
                if not row or len(row) <= max(col_c, col_d):
                    continue
                cod = str(row[col_c]).strip() if row[col_c] is not None else ""
                desc = str(row[col_d]).strip() if row[col_d] is not None else ""
                if re.match(r"^\d+$", cod):
                    resultado[cod] = desc
            tabelas[chave] = resultado
        except Exception as e:
            print(f"  [AVISO] Aba '{nome_aba}': {e}")
            tabelas[chave] = {}

    # Extrai tipoAto dos cabeçalhos da aba QualidadeDaParte (ex: "Separação (1)")
    try:
        ws_qp = wb["QualidadeDaParte"]
        tipos_ato = {}
        for row in ws_qp.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            celula = str(row[0]).strip()
            # Padrão: "Descrição (N)" ou "Descrição (N)" com número ao final
            m = re.match(r"^(.+?)\s*\((\d+)\)\s*$", celula)
            if m:
                desc_ato = m.group(1).strip()
                cod_ato = m.group(2)
                tipos_ato[cod_ato] = desc_ato
        if tipos_ato:
            tabelas["tipoAto"] = tipos_ato
    except Exception as e:
        print(f"  [AVISO] Não foi possível extrair tipoAto do XLSX: {e}")

    return tabelas


# ─────────────────────────────────────────────
# CARREGAMENTO GLOBAL DAS TABELAS
# ─────────────────────────────────────────────

TABELAS_CEP = {}
TABELAS_CESDI = {}

def _carregar_aba_xlsx(caminho: Path, nome_aba: str, col_cod: int = 0, col_desc: int = 1) -> dict:
    """Lê uma aba de XLSX e retorna {str(código): descrição}. Pula linha de cabeçalho."""
    resultado = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb[nome_aba]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or len(row) <= max(col_cod, col_desc):
                continue
            cod  = row[col_cod]
            desc = row[col_desc]
            if cod is None:
                continue
            cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
            if re.match(r"^\d+$", cod_str):
                resultado[cod_str] = str(desc).strip() if desc is not None else ""
    except Exception as e:
        print(f"  [AVISO] Não foi possível carregar aba '{nome_aba}' de {caminho.name}: {e}")
    return resultado


def _carregar_aba_xlsx_str(caminho: Path, nome_aba: str,
                            col_cod: int = 0, col_desc: int = 1) -> dict:
    """Igual a _carregar_aba_xlsx mas aceita códigos não-numéricos (ex: siglas de tribunal)."""
    resultado = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb[nome_aba]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or len(row) <= max(col_cod, col_desc):
                continue
            cod  = row[col_cod]
            desc = row[col_desc]
            if cod is None:
                continue
            cod_str = str(cod).strip()
            if cod_str and cod_str not in ("None", "nan"):
                resultado[cod_str] = str(desc).strip() if desc is not None else ""
    except Exception as e:
        print(f"  [AVISO] Não foi possível carregar aba '{nome_aba}' de {caminho.name}: {e}")
    return resultado


def _carregar_profissoes(caminho: Path, nome_aba: str = "Profissão") -> tuple:
    """
    Lê a aba Profissão com estrutura de 4 colunas:
      col 0: Código Área de Atuação
      col 1: Área de Atuação
      col 2: Código Profissão
      col 3: Profissão
    Retorna (dict_area_atuacao, dict_profissao).
    """
    area_dict = {}
    prof_dict = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        # Tenta variações do nome da aba
        aba_real = next((s for s in wb.sheetnames if "rofiss" in s), None)
        if not aba_real:
            return area_dict, prof_dict
        ws = wb[aba_real]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or len(row) < 3:
                continue
            cod_area, desc_area, cod_prof, *rest = list(row) + [None, None]
            desc_prof = rest[0] if rest else None
            if cod_area is not None:
                s = str(int(cod_area)) if isinstance(cod_area, (int, float)) else str(cod_area).strip()
                if re.match(r"^\d+$", s):
                    area_dict[s] = str(desc_area).strip() if desc_area else ""
            if cod_prof is not None:
                s = str(int(cod_prof)) if isinstance(cod_prof, (int, float)) else str(cod_prof).strip()
                if re.match(r"^\d+$", s):
                    prof_dict[s] = str(desc_prof).strip() if desc_prof else ""
    except Exception as e:
        print(f"  [AVISO] Não foi possível carregar profissões de {caminho.name}: {e}")
    return area_dict, prof_dict


def _extrair_dominio_campo(caminho: Path, nome_campo: str, col_campo: int = 0,
                            col_cod: int = 7, col_desc: int = 8) -> dict:
    """
    Extrai os códigos de domínio de um campo na aba OrientaçãoDePreenchimento.
    Encontra a linha com nome_campo e coleta pares (Código, Domínio) até o próximo campo.
    """
    resultado = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        aba = next((s for s in wb.sheetnames if "rienta" in s), None)
        if not aba:
            return resultado
        ws = wb[aba]
        capturando = False
        for row in ws.iter_rows(values_only=True):
            campo_val = row[col_campo] if len(row) > col_campo else None
            if campo_val == nome_campo:
                capturando = True
            if capturando and campo_val not in (None, nome_campo):
                break
            if capturando:
                cod  = row[col_cod]  if len(row) > col_cod  else None
                desc = row[col_desc] if len(row) > col_desc else None
                if cod is not None and desc is not None:
                    cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
                    if re.match(r"^\d+$", cod_str):
                        resultado[cod_str] = str(desc).strip()
    except Exception as e:
        print(f"  [AVISO] Não foi possível extrair domínio de '{nome_campo}' em {caminho.name}: {e}")
    return resultado


def _extrair_todos_dominios_orientacao(caminho: Path,
                                        col_campo: int = 0,
                                        col_cod: int = 7,
                                        col_desc: int = 8) -> dict:
    """
    Lê a aba OrientaçãoDePreenchimento e extrai TODOS os campos que possuem
    domínio com 2 ou mais códigos numéricos.
    Retorna {nome_campo: {código: descrição}}.
    """
    bruto: dict = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, data_only=True)
        aba = next((s for s in wb.sheetnames if "rienta" in s), None)
        if not aba:
            return {}
        ws = wb[aba]
        campo_atual = None
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) <= max(col_campo, col_cod, col_desc):
                continue
            campo_val = row[col_campo]
            cod        = row[col_cod]
            desc       = row[col_desc]
            if campo_val is not None:
                # Pega só a primeira linha do nome do campo (remove sufixos de obrigatoriedade)
                campo_atual = str(campo_val).strip().split("\n")[0].strip()
            if campo_atual and cod is not None:
                if isinstance(cod, (int, float)) and not isinstance(cod, bool):
                    cod_str = str(int(cod))
                else:
                    cod_str = str(cod).strip()
                if re.match(r"^\d+$", cod_str):
                    if campo_atual not in bruto:
                        bruto[campo_atual] = {}
                    bruto[campo_atual][cod_str] = str(desc).strip() if desc is not None else ""
    except Exception as e:
        print(f"  [AVISO] Não foi possível extrair domínios de {caminho.name}: {e}")
        return {}
    # Retorna apenas campos com 2+ códigos (domínios reais, não apenas exemplos)
    return {k: v for k, v in bruto.items() if len(v) >= 2}


def inicializar_tabelas():
    global TABELAS_CEP, TABELAS_CESDI

    # ── CEP: lê dos novos XLSX v3.3 ──────────────────────────────────────────
    api_cep    = MANUAIS_CEP_DIR / "manualAPICEP v3.3.xlsx"
    upload_cep = MANUAIS_CEP_DIR / "manualUploadCEP v3.3.xlsx"

    src_cep = api_cep if api_cep.exists() else (upload_cep if upload_cep.exists() else None)

    if src_cep:
        # ── Domínios da aba OrientaçãoDePreenchimento (CEP) ──────────────────
        dom_cep = _extrair_todos_dominios_orientacao(src_cep)
        # tipoAto: usa função específica para pegar descrições corretas (col 7/8)
        TABELAS_CEP["tipoAto"]         = _extrair_dominio_campo(src_cep, "tipoAto")
        # Outros domínios do CEP direto da aba de orientação
        TABELAS_CEP["prazoPagamento"]   = dom_cep.get("prazoPagamento", {})
        TABELAS_CEP["formaPagamento"]   = dom_cep.get("formaPagamento", {})
        TABELAS_CEP["existeBemEdireito"]= dom_cep.get("existeBemEdireito", {})
        TABELAS_CEP["acordo"]           = dom_cep.get("acordo", {})
        TABELAS_CEP["tipoDocumento"]    = dom_cep.get("tipoDocumento", {})
        TABELAS_CEP["genero"]           = dom_cep.get("sexo", {})
        TABELAS_CEP["estadoCivil"]      = dom_cep.get("estadoCivil", {})
        # nacionalidade CEP = 1=Brasileiro, 2=Estrangeiro, 3=Naturalizado (NÃO código de país)
        TABELAS_CEP["nacionalidade"]    = dom_cep.get("nacionalidade", {})
        TABELAS_CEP["capacidadeCivil"]  = dom_cep.get("capacidadeCivil", {})
        TABELAS_CEP["regimeBens"]       = dom_cep.get("regimeBens", {})
        TABELAS_CEP["tipoContato"]      = dom_cep.get("tipoContato", {})
        # tipoBemEdireito CEP = "qualificacaodeBens" na aba de orientação
        TABELAS_CEP["tipoBemEdireito"]  = dom_cep.get("qualificacaodeBens", {})
        TABELAS_CEP["tipoAtoOrigem"]    = dom_cep.get("tipoAtoOrigem", {})
        # ── Abas dedicadas ────────────────────────────────────────────────────
        for nome_aba in ("NaturezaAto", "NaturezaEscritura"):
            tab = _carregar_aba_xlsx(src_cep, nome_aba, 0, 1)
            if tab:
                TABELAS_CEP["naturezaAto"] = tab
                break
        TABELAS_CEP["naturezaLitigio"]   = _carregar_aba_xlsx(src_cep, "NaturezaLitigio", 0, 1)
        TABELAS_CEP["qualificacaoParte"]  = _carregar_aba_xlsx(src_cep, "QualidadeParte", 1, 2)
        TABELAS_CEP["pais"]              = _carregar_aba_xlsx(src_cep, "País", 0, 2)
        TABELAS_CEP["municipio"]         = _carregar_aba_xlsx(src_cep, "Município", 0, 1)
        if upload_cep.exists():
            TABELAS_CEP["estado"] = _carregar_aba_xlsx(upload_cep, "Estado", 0, 1)
        area, prof = _carregar_profissoes(src_cep)
        TABELAS_CEP["areaAtuacao"] = area
        TABELAS_CEP["profissao"]   = prof
        TABELAS_CEP["tribunal"]    = _carregar_aba_xlsx_str(src_cep, "Tribunais", 0, 1)
    else:
        TABELAS_CEP["tipoAto"]    = carregar_tabela_csv(MANUAIS_CEP_DIR / "manualAPICEP v3.csv")
        TABELAS_CEP["naturezaAto"] = carregar_tabela_csv(MANUAIS_CEP_DIR / "manualUploadCEP v3.csv")

    # ── CESDI: lê do novo XLSX API v2.5 ──────────────────────────────────────
    api_cesdi    = MANUAIS_CESDI_DIR / "manualAPICESDI v2.5.xlsx"
    upload_cesdi = MANUAIS_CESDI_DIR / "manualUploadCESDI v2.5.xlsx"

    src_cesdi = api_cesdi if api_cesdi.exists() else (upload_cesdi if upload_cesdi.exists() else None)

    if src_cesdi:
        # ── Domínios da aba OrientaçãoDePreenchimento (CESDI) ────────────────
        dom_cesdi = _extrair_todos_dominios_orientacao(src_cesdi)
        TABELAS_CESDI["prazoPagamento"]              = dom_cesdi.get("prazoPagamento", {})
        TABELAS_CESDI["formaPagamento"]              = dom_cesdi.get("formaPagamento", {})
        TABELAS_CESDI["regimeDeBensDireitosDoCasamento"] = dom_cesdi.get("regimeDeBensDireitosDoCasamento", {})
        TABELAS_CESDI["existeBemEdireitoVinculadoAoAto"] = dom_cesdi.get("existeBemEdireitoVinculadoAoAto", {})
        # tipoDocumento CESDI usa códigos IBGE (1717-1726) — mescla os dois conjuntos
        td_cesdi = {}
        td_cesdi.update(dom_cesdi.get("cpf", {}))       # códigos 1717-1721 ficam no bloco "cpf"
        td_cesdi.update(dom_cesdi.get("tipoDocumento", {}))  # 1722-1726
        TABELAS_CESDI["tipoDocumento"]    = td_cesdi
        TABELAS_CESDI["genero"]           = dom_cesdi.get("sexo", {})
        # estadoCivilParte CESDI: mescla bloco "filiacoes" (1-3) + "estadoCivilParte" (4-7)
        ec = {}
        ec.update(dom_cesdi.get("filiacoes <Lista>", {}))
        ec.update(dom_cesdi.get("estadoCivilParte", {}))
        TABELAS_CESDI["estadoCivilParte"] = ec
        # nacionalidadeParte CESDI = 1=Brasileiro, 2=Estrangeiro, 3=Naturalizado (NÃO código de país)
        TABELAS_CESDI["nacionalidadeParte"] = dom_cesdi.get("nacionalidadeParte", {})
        # capacidadeCivil CESDI: no manual o código 1=Capaz aparece na linha "profissaoParte"
        # por layout da planilha; adicionamos manualmente para garantir completude
        cc_cesdi = dict(dom_cesdi.get("capacidadeCivil", {}))
        if "1" not in cc_cesdi:
            cc_cesdi["1"] = "Capaz"
        TABELAS_CESDI["capacidadeCivil"] = cc_cesdi
        TABELAS_CESDI["tipoContatoParte"]   = dom_cesdi.get("tipoContatoParte", {})
        TABELAS_CESDI["tipoAtoOrigem"]      = dom_cesdi.get("tipoAtoOrigem", {})
        # ── Abas dedicadas CESDI ─────────────────────────────────────────────
        TABELAS_CESDI["tipoBemEdireito"]         = _carregar_aba_xlsx(src_cesdi, "TipoBensEDireito", 0, 1)
        TABELAS_CESDI["tipoReferenciaCadastral"]  = _carregar_aba_xlsx(src_cesdi, "TipoReferenciaCadastral", 0, 1)
        TABELAS_CESDI["qualidadeParte"]           = _carregar_aba_xlsx(src_cesdi, "QualidadeDaParte", 1, 2)
        TABELAS_CESDI["pais"]      = _carregar_aba_xlsx(src_cesdi, "País", 0, 2) or TABELAS_CEP.get("pais", {})
        TABELAS_CESDI["municipio"] = _carregar_aba_xlsx(src_cesdi, "Município", 0, 1) or TABELAS_CEP.get("municipio", {})
        area_c, prof_c = _carregar_profissoes(src_cesdi)
        TABELAS_CESDI["areaAtuacao"] = area_c or TABELAS_CEP.get("areaAtuacao", {})
        TABELAS_CESDI["profissao"]   = prof_c or TABELAS_CEP.get("profissao", {})
        # tipoAto CESDI: extrai dos cabeçalhos da aba QualidadeDaParte (ex: "Separação (1)")
        try:
            import openpyxl
            wb_c = openpyxl.load_workbook(src_cesdi, data_only=True)
            ws_qp = wb_c["QualidadeDaParte"]
            tipos_ato = {}
            for row in ws_qp.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                m = re.match(r"^(.+?)\s*\((\d+)\)\s*$", str(row[0]).strip())
                if m:
                    tipos_ato[m.group(2)] = m.group(1).strip()
            if tipos_ato:
                TABELAS_CESDI["tipoAto"] = tipos_ato
        except Exception as e:
            print(f"  [AVISO] Não foi possível extrair tipoAto do CESDI: {e}")
    else:
        extras = carregar_tabelas_cesdi_xlsx(upload_cesdi)
        TABELAS_CESDI.update(extras)

    # Fallback tipoAto CESDI
    if not TABELAS_CESDI.get("tipoAto"):
        TABELAS_CESDI["tipoAto"] = {
            "1": "Separação", "2": "Reconciliação",
            "3": "Conversão de Separação em Divórcio", "4": "Divórcio Direto",
            "5": "Inventário", "6": "Sobrepartilha", "7": "Rerratificação",
            "8": "Nomeação de Inventariante", "9": "Partilha",
        }

    # Fallback qualificacaoParte CEP
    if not TABELAS_CEP.get("qualificacaoParte"):
        TABELAS_CEP["qualificacaoParte"] = TABELAS_CESDI.get("qualidadeParte", {})

    # Compartilha tabelas geográficas entre sistemas se uma estiver vazia
    for chave in ("pais", "municipio", "areaAtuacao", "profissao"):
        if not TABELAS_CEP.get(chave) and TABELAS_CESDI.get(chave):
            TABELAS_CEP[chave] = TABELAS_CESDI[chave]
        if not TABELAS_CESDI.get(chave) and TABELAS_CEP.get(chave):
            TABELAS_CESDI[chave] = TABELAS_CEP[chave]


# ─────────────────────────────────────────────
# UTILITÁRIOS DE VALIDAÇÃO
# ─────────────────────────────────────────────

class Relatorio:
    def __init__(self, sistema: str, arquivo: str):
        self.sistema = sistema
        self.arquivo = arquivo
        self.erros = []
        self.avisos = []
        self.ok = []

    def erro(self, campo: str, mensagem: str):
        self.erros.append((campo, mensagem))

    def aviso(self, campo: str, mensagem: str):
        self.avisos.append((campo, mensagem))

    def sucesso(self, campo: str, mensagem: str = ""):
        self.ok.append((campo, mensagem))

    def imprimir(self):
        sep = "=" * 70
        print(f"\n{sep}")
        print(f"  RELATÓRIO DE VALIDAÇÃO SIGNO — {self.sistema}")
        print(f"  Arquivo: {self.arquivo}")
        print(f"  Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        print(sep)

        if not self.erros and not self.avisos:
            print("\n  [OK] JSON VALIDO -- Nenhum problema encontrado.\n")
        else:
            if self.erros:
                print(f"\n  [ERRO] ERROS CRITICOS ({len(self.erros)}):\n")
                for campo, msg in self.erros:
                    print(f"    [ERRO]  {campo}")
                    print(f"            -> {msg}\n")

            if self.avisos:
                print(f"\n  [AVISO] INCONSISTENCIAS ({len(self.avisos)}):\n")
                for campo, msg in self.avisos:
                    print(f"    [AVISO] {campo}")
                    print(f"            -> {msg}\n")

        print(f"  Resumo: {len(self.erros)} erro(s), {len(self.avisos)} aviso(s), "
              f"{len(self.ok)} campo(s) OK.")
        print(sep + "\n")

    @property
    def valido(self) -> bool:
        return len(self.erros) == 0


# ─────────────────────────────────────────────
# FUNÇÕES DE VALIDAÇÃO AUXILIARES
# ─────────────────────────────────────────────

def validar_data_iso(valor, campo: str, rel: Relatorio, obrigatorio: bool = False):
    """Verifica formato ISO 8601 (YYYY-MM-DD ou YYYY-MM-DDTHH:MM:SS.sssZ)."""
    if valor is None or valor == "":
        if obrigatorio:
            rel.erro(campo, "Campo obrigatório não informado.")
        return
    if not isinstance(valor, str):
        rel.erro(campo, f"Esperado string com data ISO 8601, recebido {type(valor).__name__}: {valor!r}")
        return
    # Aceita YYYY-MM-DD ou YYYY-MM-DDTHH:MM:SS... 
    padrao = r"^\d{4}-\d{2}-\d{2}(T\d{2}:\d{2}:\d{2}(\.\d+)?(Z|[+-]\d{2}:\d{2})?)?$"
    if not re.match(padrao, valor):
        rel.erro(campo, f"Formato de data inválido: {valor!r}. Use ISO 8601 (ex: 2024-06-01 ou 2024-06-01T00:00:00.000Z).")
    else:
        rel.sucesso(campo, f"Data OK: {valor}")


def validar_string_campo_obrigatorio(valor, campo: str, rel: Relatorio):
    if valor is None or (isinstance(valor, str) and valor.strip() == ""):
        rel.erro(campo, "Campo string obrigatório está ausente ou vazio.")
    elif not isinstance(valor, str):
        rel.erro(campo, f"Deve ser uma string, mas recebeu {type(valor).__name__}: {valor!r}")
    else:
        rel.sucesso(campo)


def validar_inteiro(valor, campo: str, rel: Relatorio, obrigatorio: bool = True,
                     valores_validos: dict = None, nome_tabela: str = ""):
    """Valida que o campo é inteiro e, opcionalmente, cruza com tabela de códigos."""
    if valor is None:
        if obrigatorio:
            rel.erro(campo, "Campo inteiro obrigatório não informado.")
        return
    if not isinstance(valor, int) or isinstance(valor, bool):
        rel.erro(campo, f"Tipo inválido: esperado inteiro, recebido {type(valor).__name__}: {valor!r}")
        return
    if valores_validos is not None:
        chave = str(valor)
        if chave not in valores_validos:
            codigos_disponiveis = ", ".join(
                f"{k}={v}" for k, v in sorted(valores_validos.items(), key=lambda x: int(x[0]))
            )
            rel.erro(campo,
                     f"Código {valor} não reconhecido na tabela '{nome_tabela}'.\n"
                     f"            Códigos válidos: {codigos_disponiveis}")
        else:
            rel.sucesso(campo, f"Código {valor} → {valores_validos[chave]}")


def validar_string_numerica(valor, campo: str, rel: Relatorio,
                             comprimento: int = None, nome: str = ""):
    """
    Valida CPF, CNPJ, CNS, CEP — devem ser strings preservando zeros à esquerda.
    """
    if valor is None or valor == "":
        return  # campo opcional; obrigatoriedade checada separadamente
    if not isinstance(valor, str):
        rel.erro(campo,
                 f"{nome or campo} deve ser string (para preservar zeros à esquerda), "
                 f"mas recebeu {type(valor).__name__}: {valor!r}")
        return
    # Remove formatação para contar dígitos
    apenas_digitos = re.sub(r"[.\-/]", "", valor).strip()
    if comprimento and len(apenas_digitos) != comprimento:
        rel.aviso(campo,
                  f"{nome or campo} tem {len(apenas_digitos)} dígito(s); "
                  f"esperado {comprimento}. Valor: {valor!r}")
    else:
        rel.sucesso(campo, f"{nome or campo} OK: {valor}")


def campo_presente(dados: dict, campo: str) -> bool:
    return campo in dados and dados[campo] is not None


# ─────────────────────────────────────────────
# VALIDAÇÃO CEP
# ─────────────────────────────────────────────

def validar_cep_json(dados: dict, rel: Relatorio):
    tabelas = TABELAS_CEP

    # ── Campos raiz obrigatórios ──────────────────────────────────
    campos_int_obrig = {
        "tipoAto":      ("tipoAto",     tabelas.get("tipoAto", {})),
        "status":       ("status",      {}),
        "livroInicial": ("livroInicial",{}),
        "folhaInicial": ("folhaInicial",{}),
        "folhaFinal":   ("folhaFinal",  {}),
    }
    for campo, (nome_tab, tab) in campos_int_obrig.items():
        validar_inteiro(dados.get(campo), campo, rel, obrigatorio=True,
                        valores_validos=tab if tab else None,
                        nome_tabela=nome_tab if tab else "")

    # naturezaAto: obrigatório SOMENTE quando tipoAto = 1 (Escritura)
    tipo_ato = dados.get("tipoAto")
    natureza = dados.get("naturezaAto")
    if tipo_ato == 1:
        validar_inteiro(natureza, "naturezaAto", rel, obrigatorio=True,
                        valores_validos=tabelas.get("naturezaAto"),
                        nome_tabela="naturezaAto")
    elif natureza is not None:
        # Se informado em outro tipo de ato, apenas valida o código se existir na tabela
        validar_inteiro(natureza, "naturezaAto", rel, obrigatorio=False,
                        valores_validos=tabelas.get("naturezaAto"),
                        nome_tabela="naturezaAto")

    # ── Campos inteiros opcionais SEM domínio específico ─────────────────────
    for campo in ["tipoInvalidacaoAto", "livroFinal", "valorOperacao"]:
        v = dados.get(campo)
        if v is not None and not isinstance(v, (int, float)) or isinstance(v, bool):
            if campo in dados and isinstance(dados[campo], str) and not dados[campo].isdigit():
                rel.aviso(campo, f"Esperado número, recebeu string: {dados[campo]!r}")

    # ── Campos inteiros opcionais COM domínio ────────────────────────────────
    campos_com_dom_raiz = [
        ("prazoPagamento",    "prazoPagamento",    "Prazo de Pagamento"),
        ("formaPagamento",    "formaPagamento",    "Forma de Pagamento"),
        ("existeBemEdireito", "existeBemEdireito", "Existência de Bem/Direito"),
        ("acordo",            "acordo",            "Acordo"),
        ("naturezaLitigio",   "naturezaLitigio",   "Natureza do Litígio"),
    ]
    for campo, chave_tab, nome_tab in campos_com_dom_raiz:
        v = dados.get(campo)
        if v is not None:
            tab = tabelas.get(chave_tab)
            validar_inteiro(v, campo, rel, obrigatorio=False,
                            valores_validos=tab if tab else None,
                            nome_tabela=nome_tab if tab else "")

    # ── Campos de data ────────────────────────────────────────────
    validar_data_iso(dados.get("dataAto"),      "dataAto",      rel, obrigatorio=True)
    validar_data_iso(dados.get("dataContrato"), "dataContrato", rel)
    validar_data_iso(dados.get("dataInclusao"), "dataInclusao", rel)

    # ── Campo reservaDePoderes (booleano) ─────────────────────────
    if "reservaDePoderes" in dados:
        if not isinstance(dados["reservaDePoderes"], bool):
            rel.aviso("reservaDePoderes", f"Esperado booleano, recebeu: {dados['reservaDePoderes']!r}")

    # ── Partes ────────────────────────────────────────────────────
    partes = dados.get("partes", [])
    if not isinstance(partes, list):
        rel.erro("partes", "Campo 'partes' deve ser uma lista.")
    elif len(partes) == 0:
        rel.aviso("partes", "Nenhuma parte informada no ato.")
    else:
        for i, parte in enumerate(partes):
            prefixo = f"partes[{i}]"
            _validar_parte_cep(parte, i, prefixo, rel)

    # ── Bens e Direitos ───────────────────────────────────────────
    bens = dados.get("bensEdireitos", [])
    existe_bem = dados.get("existeBemEdireito")
    if isinstance(bens, list):
        # existeBemEdireito=1 exige que a lista tenha pelo menos um bem preenchido
        if existe_bem == 1 and len(bens) == 0:
            rel.erro("bensEdireitos",
                     "Campo 'existeBemEdireito' indica que há bens/direitos (valor=1), "
                     "mas a lista 'bensEdireitos' está vazia.")
        for i, bem in enumerate(bens):
            if not isinstance(bem, dict) or not bem:
                rel.erro(f"bensEdireitos[{i}]",
                         "Item de bem e direito vazio ou inválido. "
                         "Preencha os dados ou remova o item da lista.")
            else:
                _validar_bem_cep(bem, i, f"bensEdireitos[{i}]", rel)

    # ── Atos de Origem ───────────────────────────────────────────
    atos_origem = dados.get("atosOrigem", [])
    if isinstance(atos_origem, list):
        for i, ao in enumerate(atos_origem):
            _validar_ato_origem_cep(ao, i, f"atosOrigem[{i}]", rel)

    # ── Attachments ───────────────────────────────────────────────
    attachments = dados.get("attachments", [])
    if isinstance(attachments, list):
        for i, att in enumerate(attachments):
            prefixo = f"attachments[{i}]"
            for campo in ["base64Content", "extensaoArquivo", "nome"]:
                if not att.get(campo):
                    rel.aviso(f"{prefixo}.{campo}", f"Campo de anexo vazio ou ausente.")


def _validar_parte_cep(parte: dict, idx: int, prefixo: str, rel: Relatorio):
    tabelas = TABELAS_CEP

    # qualificacaoParte obrigatório
    validar_inteiro(parte.get("qualificacaoParte"), f"{prefixo}.qualificacaoParte", rel,
                    obrigatorio=True,
                    valores_validos=tabelas.get("qualificacaoParte"),
                    nome_tabela="qualificacaoParte")

    # nomeParte obrigatório
    validar_string_campo_obrigatorio(parte.get("nomeParte"), f"{prefixo}.nomeParte", rel)

    # CPF / CNPJ: deve ser string
    validar_string_numerica(parte.get("cpf"), f"{prefixo}.cpf", rel, nome="CPF/CNPJ")

    # cpfConjuge: deve ser string
    if parte.get("cpfConjuge"):
        validar_string_numerica(parte["cpfConjuge"], f"{prefixo}.cpfConjuge", rel,
                                comprimento=11, nome="CPF do Cônjuge")

    # CEP: deve ser string
    if parte.get("cep"):
        validar_string_numerica(parte["cep"], f"{prefixo}.cep", rel,
                                comprimento=8, nome="CEP")

    # datas
    for campo_data in ["dataCasamento", "dataEmissao", "dataNascimento", "dataObito"]:
        if parte.get(campo_data):
            validar_data_iso(parte[campo_data], f"{prefixo}.{campo_data}", rel)

    # inteiros opcionais com validação de domínio (tabelas dos manuais)
    campos_com_dominio = [
        ("tipoDocumento",  "tipoDocumento",  "Tipo de Documento"),
        ("genero",         "genero",         "Gênero"),
        ("estadoCivil",    "estadoCivil",    "Estado Civil"),
        ("nacionalidade",  "nacionalidade",  "Nacionalidade"),
        ("capacidadeCivil","capacidadeCivil","Capacidade Civil"),
        ("regimeBens",     "regimeBens",     "Regime de Bens"),
        ("tipoContato",    "tipoContato",    "Tipo de Contato"),
        ("municipio",      "municipio",      "Município"),
        ("paisOrigem",     "pais",           "País de Origem"),
        ("areaAtuacao",    "areaAtuacao",    "Área de Atuação"),
        ("profissao",      "profissao",      "Profissão"),
    ]
    for campo, chave_tab, nome_tab in campos_com_dominio:
        v = parte.get(campo)
        if v is not None:
            tab = tabelas.get(chave_tab)
            validar_inteiro(v, f"{prefixo}.{campo}", rel, obrigatorio=False,
                            valores_validos=tab if tab else None,
                            nome_tabela=nome_tab if tab else "")

    # semFiliacoes: booleano
    if "semFiliacoes" in parte:
        if not isinstance(parte["semFiliacoes"], bool):
            rel.aviso(f"{prefixo}.semFiliacoes", f"Esperado booleano, recebido: {parte['semFiliacoes']!r}")

    # filiacoes: lista de strings
    filiacoes = parte.get("filiacoes", [])
    if filiacoes and not isinstance(filiacoes, list):
        rel.erro(f"{prefixo}.filiacoes", "Campo 'filiacoes' deve ser uma lista de strings.")


def _validar_bem_cep(bem: dict, idx: int, prefixo: str, rel: Relatorio):
    tabelas = TABELAS_CEP

    # qualificacaodeBens: obrigatório (manual CEP diz SIM)
    validar_inteiro(bem.get("qualificacaodeBens"), f"{prefixo}.qualificacaodeBens", rel,
                    obrigatorio=True,
                    valores_validos=tabelas.get("tipoBemEdireito"),
                    nome_tabela="qualificacaodeBens")

    # cep do imóvel: string
    if bem.get("cep"):
        validar_string_numerica(bem["cep"], f"{prefixo}.cep", rel, comprimento=8, nome="CEP do imóvel")

    # cin: string (Código Imobiliário Nacional)
    if bem.get("cin") and not isinstance(bem["cin"], str):
        rel.erro(f"{prefixo}.cin", f"CIN deve ser string, recebido: {type(bem['cin']).__name__}")

    # valor: numérico
    for campo_num in ["valor", "valorFiscal", "valorImovel",
                       "quantidadeAreaConstruida", "quantidadeAreaTotal"]:
        v = bem.get(campo_num)
        if v is not None and not isinstance(v, (int, float)):
            rel.aviso(f"{prefixo}.{campo_num}",
                      f"Esperado número, recebido {type(v).__name__}: {v!r}")

    # titulares: obrigatório e não pode ser vazio (manual diz SIM)
    titulares = bem.get("titulares")
    if titulares is None or titulares == []:
        rel.erro(f"{prefixo}.titulares",
                 "Lista de titulares obrigatória e não pode ser vazia.")
    elif isinstance(titulares, list):
        for j, tit in enumerate(titulares):
            if not isinstance(tit, dict) or not tit:
                rel.erro(f"{prefixo}.titulares[{j}]", "Titular vazio ou inválido.")
            elif tit.get("cpf") and not isinstance(tit["cpf"], str):
                rel.erro(f"{prefixo}.titulares[{j}].cpf",
                         "CPF do titular deve ser string (preserva zeros à esquerda).")


def _validar_ato_origem_cep(ao: dict, idx: int, prefixo: str, rel: Relatorio):
    # CNS: string
    if ao.get("numeroCns") and not isinstance(ao["numeroCns"], str):
        rel.erro(f"{prefixo}.numeroCns", "CNS deve ser string (preserva zeros à esquerda).")

    # tribunal: string — valida contra tabela de tribunais
    if ao.get("tribunal"):
        cod_trib = str(ao["tribunal"]).strip()
        tab_trib = TABELAS_CEP.get("tribunal", {})
        if tab_trib and cod_trib not in tab_trib:
            codigos = ", ".join(sorted(tab_trib.keys()))
            rel.aviso(f"{prefixo}.tribunal",
                      f"Código de tribunal '{cod_trib}' não reconhecido.\n"
                      f"            Códigos válidos: {codigos}")
        elif tab_trib:
            rel.sucesso(f"{prefixo}.tribunal", f"{cod_trib} → {tab_trib[cod_trib]}")

    # municipioCartorioAtoOrigem: valida contra tabela de municípios
    v_mun = ao.get("municipioCartorioAtoOrigem")
    if v_mun is not None:
        tab_mun = TABELAS_CEP.get("municipio")
        validar_inteiro(v_mun, f"{prefixo}.municipioCartorioAtoOrigem", rel, obrigatorio=False,
                        valores_validos=tab_mun if tab_mun else None,
                        nome_tabela="Município (IBGE)" if tab_mun else "")

    # booleans
    for campo_bool in ["atosAnteriores", "cartorioAtual", "outroCartorio"]:
        v = ao.get(campo_bool)
        if v is not None and not isinstance(v, bool):
            rel.aviso(f"{prefixo}.{campo_bool}", f"Esperado booleano, recebido: {v!r}")


# ─────────────────────────────────────────────
# VALIDAÇÃO CESDI
# ─────────────────────────────────────────────

def validar_cesdi_json(dados: dict, rel: Relatorio):
    tabelas = TABELAS_CESDI

    # ── Campos raiz obrigatórios ──────────────────────────────────
    validar_inteiro(dados.get("tipoAto"), "tipoAto", rel, obrigatorio=True,
                    valores_validos=tabelas.get("tipoAto", {}),
                    nome_tabela="tipoAto CESDI")

    validar_inteiro(dados.get("statusAto"), "statusAto", rel, obrigatorio=True)

    validar_data_iso(dados.get("dataAto"), "dataAto", rel, obrigatorio=True)

    for campo in ["livroInicial", "folhaInicial", "folhaFinal"]:
        validar_inteiro(dados.get(campo), campo, rel, obrigatorio=True)

    # ── Campos de data opcionais ──────────────────────────────────
    if dados.get("dataContrato"):
        validar_data_iso(dados["dataContrato"], "dataContrato", rel)

    # ── Validações condicionais ao tipoAto ────────────────────────
    # Atos de casamento (1=Separação, 2=Reconciliação, 3=Conversão, 4=Divórcio Direto):
    # os campos abaixo são obrigatórios conforme o manual CESDI API
    TIPOS_CASAMENTO = {1, 2, 3, 4}
    tipo_ato_cesdi = dados.get("tipoAto")
    eh_ato_casamento = isinstance(tipo_ato_cesdi, int) and tipo_ato_cesdi in TIPOS_CASAMENTO

    validar_inteiro(dados.get("regimeDeBensDireitosDoCasamento"),
                    "regimeDeBensDireitosDoCasamento", rel,
                    obrigatorio=eh_ato_casamento,
                    valores_validos=tabelas.get("regimeDeBensDireitosDoCasamento"),
                    nome_tabela="Regime de Bens do Casamento")

    validar_data_iso(dados.get("dataCasamento"), "dataCasamento", rel,
                     obrigatorio=eh_ato_casamento)

    for campo in ["quantidadefilhosMaiores", "quantidadefilhosMenores"]:
        v = dados.get(campo)
        if eh_ato_casamento and v is None:
            rel.erro(campo, "Campo inteiro obrigatório não informado.")
        elif v is not None and not isinstance(v, int):
            rel.aviso(campo, f"Esperado inteiro, recebido {type(v).__name__}: {v!r}")

    # ── Campos inteiros opcionais sem domínio ─────────────────────
    v = dados.get("livroFinal")
    if v is not None and not isinstance(v, int):
        rel.aviso("livroFinal", f"Esperado inteiro, recebido {type(v).__name__}: {v!r}")

    # ── Campos inteiros opcionais com domínio ─────────────────────
    campos_dom_cesdi = [
        ("prazoPagamento",    "prazoPagamento",    "Prazo de Pagamento"),
        ("formaPagamento",    "formaPagamento",    "Forma de Pagamento"),
        ("municipioCartorioAtoOrigem", "municipio", "Município"),
        ("tipoAtoOrigem", "tipoAtoOrigem", "Tipo do Ato de Origem"),
    ]
    for campo, chave_tab, nome_tab in campos_dom_cesdi:
        v = dados.get(campo)
        if v is not None:
            tab = tabelas.get(chave_tab)
            validar_inteiro(v, campo, rel, obrigatorio=False,
                            valores_validos=tab if tab else None,
                            nome_tabela=nome_tab if tab else "")

    # ── existeBemEdireitoVinculadoAoAto: string ("0" ou "1") ─────
    v_bem = dados.get("existeBemEdireitoVinculadoAoAto")
    if v_bem is not None and not isinstance(v_bem, str):
        rel.aviso("existeBemEdireitoVinculadoAoAto",
                  f"Esperado string, recebido {type(v_bem).__name__}: {v_bem!r}")

    # ── Booleans ──────────────────────────────────────────────────
    for campo_bool in ["cartorioAtual", "cartorioNaoCadastrado"]:
        v = dados.get(campo_bool)
        if v is not None and not isinstance(v, bool):
            rel.aviso(campo_bool, f"Esperado booleano, recebido: {v!r}")

    # ── cartorio (CNS): obrigatório exceto quando cartorioNaoCadastrado=true ─
    cartorio_nao_cad = dados.get("cartorioNaoCadastrado")
    cartorio_val = dados.get("cartorio")
    if cartorio_nao_cad is not True:
        if cartorio_val is not None and not isinstance(cartorio_val, str):
            rel.erro("cartorio", "CNS do cartório deve ser string.")
        elif (cartorio_val is None or (isinstance(cartorio_val, str) and cartorio_val.strip() == "")):
            if dados.get("cartorioAtual") is not True:
                rel.aviso("cartorio",
                          "CNS do cartório de origem não informado. "
                          "Preencha ou indique cartorioNaoCadastrado=true.")

    # ── Partes ────────────────────────────────────────────────────
    partes = dados.get("partes", [])
    if not isinstance(partes, list):
        rel.erro("partes", "Campo 'partes' deve ser uma lista.")
    elif len(partes) == 0:
        rel.aviso("partes", "Nenhuma parte informada no ato.")
    else:
        for i, parte in enumerate(partes):
            _validar_parte_cesdi(parte, i, f"partes[{i}]", rel, tipo_ato_cesdi)

    # ── Bens e Direitos ───────────────────────────────────────────
    bens = dados.get("bensEdireitos", [])
    existe_bem_cesdi = dados.get("existeBemEdireitoVinculadoAoAto")
    if isinstance(bens, list):
        if existe_bem_cesdi == "1" and len(bens) == 0:
            rel.erro("bensEdireitos",
                     "Campo 'existeBemEdireitoVinculadoAoAto' indica que há bens/direitos, "
                     "mas a lista 'bensEdireitos' está vazia.")
        for i, bem in enumerate(bens):
            if not isinstance(bem, dict) or not bem:
                rel.erro(f"bensEdireitos[{i}]",
                         "Item de bem e direito vazio ou inválido. "
                         "Preencha os dados ou remova o item da lista.")
            else:
                _validar_bem_cesdi(bem, i, f"bensEdireitos[{i}]", rel)

    # ── Anexos ───────────────────────────────────────────────────
    for lista_anexos in ["anexos", "anexosEspecificos"]:
        for i, att in enumerate(dados.get(lista_anexos, [])):
            prefixo = f"{lista_anexos}[{i}]"
            for campo in ["base64Content", "extensaoArquivo", "nome"]:
                if not att.get(campo):
                    rel.aviso(f"{prefixo}.{campo}", "Campo de anexo vazio ou ausente.")


def _validar_parte_cesdi(parte: dict, idx: int, prefixo: str, rel: Relatorio,
                         tipo_ato: int | None = None):
    tabelas = TABELAS_CESDI
    TIPOS_CASAMENTO = {1, 2, 3, 4}
    eh_ato_casamento = isinstance(tipo_ato, int) and tipo_ato in TIPOS_CASAMENTO

    # qualificacaoParte
    validar_inteiro(parte.get("qualificacaoParte"), f"{prefixo}.qualificacaoParte", rel,
                    obrigatorio=True,
                    valores_validos=tabelas.get("qualidadeParte"),
                    nome_tabela="QualidadeDaParte (XLSX)")

    # nomeParte obrigatório
    validar_string_campo_obrigatorio(parte.get("nomeParte"), f"{prefixo}.nomeParte", rel)

    # CPF: deve ser string
    validar_string_numerica(parte.get("cpf"), f"{prefixo}.cpf", rel, nome="CPF/CNPJ")

    # CEP: deve ser string
    if parte.get("cepResidencia"):
        validar_string_numerica(parte["cepResidencia"], f"{prefixo}.cepResidencia", rel,
                                comprimento=8, nome="CEP de residência")

    # cpfConjuge: string
    if parte.get("cpfConjuge"):
        validar_string_numerica(parte["cpfConjuge"], f"{prefixo}.cpfConjuge", rel,
                                comprimento=11, nome="CPF do Cônjuge")

    # datas
    for campo_data in ["dataEmissaoDocumento", "dataNascimentoParte",
                       "dataCasamento", "dataObito"]:
        if parte.get(campo_data):
            validar_data_iso(parte[campo_data], f"{prefixo}.{campo_data}", rel)

    # inteiros com validação de domínio (tabelas dos manuais)
    campos_com_dominio = [
        ("tipoDocumento",     "tipoDocumento",      "Tipo de Documento"),
        ("genero",            "genero",             "Gênero"),
        ("estadoCivilParte",  "estadoCivilParte",   "Estado Civil"),
        ("capacidadeCivil",   "capacidadeCivil",    "Capacidade Civil"),
        ("tipoContatoParte",  "tipoContatoParte",   "Tipo de Contato"),
        ("regimeDeBensDireitosDoCasamento", "regimeDeBensDireitosDoCasamento", "Regime de Bens do Casamento"),
        ("cidadeResidencia",    "municipio",          "Município"),
        ("codigoPaisParte",     "pais",               "País da Parte"),
        ("codigoPaisResidencia","pais",               "País de Residência"),
        # nacionalidadeParte CESDI = 1=Brasileiro, 2=Estrangeiro, 3=Naturalizado
        ("nacionalidadeParte",  "nacionalidadeParte", "Nacionalidade"),
        # areaAtuacaoParte e profissaoParte no CESDI são campos texto livre (não código)
    ]
    for campo, chave_tab, nome_tab in campos_com_dominio:
        v = parte.get(campo)
        if v is not None:
            tab = tabelas.get(chave_tab)
            validar_inteiro(v, f"{prefixo}.{campo}", rel, obrigatorio=False,
                            valores_validos=tab if tab else None,
                            nome_tabela=nome_tab if tab else "")

    # responsavelFilhosMenores: obrigatório para atos de separação/divórcio (tipoAto 1-4)
    v_resp = parte.get("responsavelFilhosMenores")
    if eh_ato_casamento and v_resp is None:
        rel.erro(f"{prefixo}.responsavelFilhosMenores",
                 "Campo obrigatório para atos de separação/divórcio (tipoAto 1-4).")
    elif v_resp is not None and not isinstance(v_resp, bool):
        rel.aviso(f"{prefixo}.responsavelFilhosMenores",
                  f"Esperado booleano, recebido: {v_resp!r}")

    # naoPossuiFiliacao: boolean opcional
    v_fil = parte.get("naoPossuiFiliacao")
    if v_fil is not None and not isinstance(v_fil, bool):
        rel.aviso(f"{prefixo}.naoPossuiFiliacao", f"Esperado booleano, recebido: {v_fil!r}")

    # filiacoes: lista
    filiacoes = parte.get("filiacoes", [])
    if filiacoes and not isinstance(filiacoes, list):
        rel.erro(f"{prefixo}.filiacoes", "Campo 'filiacoes' deve ser lista de strings.")


def _validar_bem_cesdi(bem: dict, idx: int, prefixo: str, rel: Relatorio):
    tabelas = TABELAS_CESDI

    # tipoBemEdireito: obrigatório (manual CESDI diz SIM)
    validar_inteiro(bem.get("tipoBemEdireito"), f"{prefixo}.tipoBemEdireito", rel,
                    obrigatorio=True,
                    valores_validos=tabelas.get("tipoBemEdireito"),
                    nome_tabela="TipoBensEDireito (XLSX)")

    # tipoReferenciaCadastral
    if "referenciaCadastralImovel" in bem:
        validar_inteiro(bem.get("referenciaCadastralImovel"),
                        f"{prefixo}.referenciaCadastralImovel", rel,
                        obrigatorio=False,
                        valores_validos=tabelas.get("tipoReferenciaCadastral"),
                        nome_tabela="TipoReferenciaCadastral (XLSX)")

    # CEP do imóvel: string
    if bem.get("cep"):
        validar_string_numerica(bem["cep"], f"{prefixo}.cep", rel,
                                comprimento=8, nome="CEP do imóvel")

    # CIN: string
    if bem.get("cin") and not isinstance(bem["cin"], str):
        rel.erro(f"{prefixo}.cin", "CIN deve ser string.")

    # numéricos
    for campo_num in ["valorDoBem", "valorImovel", "valorFiscal",
                       "quantidadeAreaTotal", "quantidadeUnidadeAreaConstruida"]:
        v = bem.get(campo_num)
        if v is not None and not isinstance(v, (int, float)):
            rel.aviso(f"{prefixo}.{campo_num}",
                      f"Esperado número, recebido {type(v).__name__}: {v!r}")

    # titulares: obrigatório e não pode ser vazio (manual diz SIM)
    titulares = bem.get("titulares")
    if titulares is None or titulares == []:
        rel.erro(f"{prefixo}.titulares",
                 "Lista de titulares obrigatória e não pode ser vazia.")
    elif isinstance(titulares, list):
        for j, tit in enumerate(titulares):
            if not isinstance(tit, dict) or not tit:
                rel.erro(f"{prefixo}.titulares[{j}]", "Titular vazio ou inválido.")
            elif tit.get("cpf") and not isinstance(tit["cpf"], str):
                rel.erro(f"{prefixo}.titulares[{j}].cpf",
                         "CPF do titular deve ser string (preserva zeros à esquerda).")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Validador local de JSON para sistemas CEP e CESDI da SIGNO."
    )
    parser.add_argument("arquivo", help="Caminho para o arquivo JSON a ser validado.")
    parser.add_argument(
        "--sistema", "-s",
        choices=["CEP", "CESDI"],
        required=True,
        help="Sistema alvo: CEP ou CESDI."
    )
    args = parser.parse_args()

    # Verifica existência do arquivo
    caminho_json = Path(args.arquivo)
    if not caminho_json.exists():
        print(f"\n[ERRO] Arquivo não encontrado: {caminho_json}")
        sys.exit(1)

    # Lê JSON
    try:
        with open(caminho_json, encoding="utf-8") as f:
            dados = json.load(f)
    except json.JSONDecodeError as e:
        print(f"\n[ERRO FATAL] JSON inválido (erro de sintaxe): {e}")
        sys.exit(1)

    # Carrega tabelas dos manuais
    print(f"\nCarregando tabelas dos manuais...")
    inicializar_tabelas()
    print("  OK — Tabelas carregadas.")

    # Valida
    rel = Relatorio(sistema=args.sistema, arquivo=str(caminho_json))

    if args.sistema == "CEP":
        validar_cep_json(dados, rel)
    else:
        validar_cesdi_json(dados, rel)

    rel.imprimir()
    sys.exit(0 if rel.valido else 1)


if __name__ == "__main__":
    main()
